#!/usr/bin/env python3
"""
AI Powered Security Camera (Jetson Nano - FIXED VERSION)
- Attendance for known faces (Excel)
- Unknown face: photo alert (Telegram)
- Objects/Animals (non-person): photo alert (per label, 1/min)
- Motion-only (no faces/objects): text alert (30s)

FIXES:
- Removed ultralytics dependency (use torch.hub instead)
- Added proper error handling
- Threading for Telegram (non-blocking)
- Frame skipping for performance
- Resource cleanup
"""

import os
import cv2
import time
import torch
import telepot
import openpyxl
import datetime
import numpy as np
import face_recognition
import threading
from pathlib import Path
from queue import Queue
from flask import Flask, render_template_string, Response
import io
import base64

############################################
# ============ CONFIGURATION ==============
############################################
# Paths
known_faces_dir = "known_faces"
unknown_faces_dir = "unknown_faces"
attendance_file = "attendance.xlsx"

# Telegram
TELEGRAM_TOKEN = "8271271547:AAGzGKDc8woVP8w4FC_LhJQLLCLFcqd2SQA"   # <-- REPLACE THIS
CHAT_IDS = ["1668233349"]        # <-- REPLACE THIS

# Cooldowns (seconds)
cooldown_unknown = 30
cooldown_motion  = 30
cooldown_object  = 60  # per object type

# Detection thresholds
conf_thres = 0.5
iou_thres  = 0.45
min_motion_area = 600

# Camera source
# For Jetson Nano with USB camera:
# - Run: ls /dev/video* to see available cameras
# - CSI camera (built-in ribbon): typically /dev/video0
# - USB camera: typically /dev/video1, /dev/video2, etc.
# - Set camera_index to the USB camera index (usually 1)
camera_index = 1  # Change to 0 for CSI, or 1+ for USB

# Performance settings
PROCESS_EVERY_N_FRAMES = 3  # Process every 3rd frame for performance
FACE_SCALE_FACTOR = 0.25    # Downscale for face detection

# Web streaming
WEB_PORT = 10000

############################################
# ============ FLASK SETUP ================
############################################
app = Flask(__name__)
current_frame = None
frame_lock = threading.Lock()

HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>AI Security Camera - Jetson Nano</title>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        body { font-family: Arial; text-align: center; background: #222; color: #fff; }
        h1 { color: #0f0; }
        img { max-width: 90%; height: auto; border: 2px solid #0f0; }
        .status { color: #0f0; font-size: 18px; margin: 10px 0; }
    </style>
</head>
<body>
    <h1>🎥 AI Security Camera - Live Feed</h1>
    <p class="status">Streaming from Jetson Nano</p>
    <img src="/video_feed" width="800">
</body>
</html>
"""

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/video_feed')
def video_feed():
    """Stream video frames as MJPEG"""
    def generate():
        while True:
            with frame_lock:
                if current_frame is None:
                    continue
                ret, buffer = cv2.imencode('.jpg', current_frame)
                frame_data = buffer.tobytes()
            
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n'
                   b'Content-Length: ' + f'{len(frame_data)}'.encode() + b'\r\n\r\n'
                   + frame_data + b'\r\n')
            time.sleep(0.03)  # ~30 FPS
    
    return Response(generate(),
                    mimetype='multipart/x-mixed-replace; boundary=frame')

############################################
# ============ PREP FOLDERS ===============
############################################
try:
    os.makedirs(unknown_faces_dir, exist_ok=True)
    os.makedirs("object_alerts", exist_ok=True)
    print("[INFO] Directories created/verified")
except Exception as e:
    print(f"[ERROR] Could not create directories: {e}")
    exit(1)

############################################
# ========== TELEGRAM SETUP ===============
############################################
telegram_queue = Queue()
telegram_enabled = False

try:
    bot = telepot.Bot(TELEGRAM_TOKEN)
    # Test connection
    bot.getMe()
    telegram_enabled = True
    print("[INFO] Telegram bot connected successfully")
except Exception as e:
    print(f"[WARN] Telegram not configured properly: {e}")
    print("[WARN] Continuing without Telegram alerts")

def telegram_worker():
    """Background thread for sending Telegram messages"""
    while True:
        try:
            item = telegram_queue.get()
            if item is None:
                break
            
            msg_type, data = item
            
            for chat_id in CHAT_IDS:
                try:
                    if msg_type == "text":
                        bot.sendMessage(chat_id, data)
                    elif msg_type == "photo":
                        msg, photo_path = data
                        bot.sendMessage(chat_id, msg)
                        if os.path.exists(photo_path):
                            with open(photo_path, "rb") as photo:
                                bot.sendPhoto(chat_id, photo=photo)
                except Exception as e:
                    print(f"[ERROR] Telegram send failed for {chat_id}: {e}")
            
            telegram_queue.task_done()
        except Exception as e:
            print(f"[ERROR] Telegram worker error: {e}")

# Start Telegram worker thread
if telegram_enabled:
    telegram_thread = threading.Thread(target=telegram_worker, daemon=True)
    telegram_thread.start()

def send_telegram_alert(msg_type, data):
    """Queue a Telegram alert"""
    if telegram_enabled:
        telegram_queue.put((msg_type, data))

############################################
# ======== LOAD KNOWN FACES DB ============
############################################
known_face_encodings = []
known_face_names = []

print("[INFO] Loading known faces...")
if os.path.exists(known_faces_dir):
    for filename in os.listdir(known_faces_dir):
        if filename.lower().endswith((".jpg", ".jpeg", ".png")):
            path = os.path.join(known_faces_dir, filename)
            try:
                image = face_recognition.load_image_file(path)
                encs = face_recognition.face_encodings(image)
                if len(encs) > 0:
                    known_face_encodings.append(encs[0])
                    known_face_names.append(os.path.splitext(filename)[0])
                    print(f"[INFO] Loaded: {filename}")
                else:
                    print(f"[WARN] No face detected in: {filename}")
            except Exception as e:
                print(f"[ERROR] Failed to load {filename}: {e}")
else:
    print(f"[WARN] Known faces directory not found: {known_faces_dir}")
    os.makedirs(known_faces_dir, exist_ok=True)
    print(f"[INFO] Created directory. Add face images to: {known_faces_dir}")

print(f"[INFO] Total known faces loaded: {len(known_face_encodings)}")

############################################
# ============ EXCEL SETUP ================
############################################
def init_excel():
    """Initialize Excel file"""
    try:
        if not os.path.exists(attendance_file):
            wb = openpyxl.Workbook()
            sh = wb.active
            sh.title = "Attendance"
            sh.append(["Name", "Date", "Time"])
            wb.save(attendance_file)
            print(f"[INFO] Created attendance file: {attendance_file}")
    except Exception as e:
        print(f"[ERROR] Could not create Excel file: {e}")

def mark_attendance(name):
    """Mark attendance in Excel"""
    try:
        wb = openpyxl.load_workbook(attendance_file)
        sh = wb["Attendance"] if "Attendance" in wb.sheetnames else wb.active
        
        now = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        
        # Check for duplicate entry today
        for row in sh.iter_rows(min_row=2, values_only=True):
            if row[0] == name and row[1] == date_str:
                wb.close()
                return False
        
        sh.append([name, date_str, time_str])
        wb.save(attendance_file)
        wb.close()
        print(f"[INFO] Attendance marked: {name} @ {time_str}")
        return True
    except Exception as e:
        print(f"[ERROR] Failed to mark attendance: {e}")
        return False

init_excel()

############################################
# ======== YOLOV5 LOADING (FIXED) =========
############################################
print("[INFO] Loading YOLOv5 model...")
model = None

try:
    # Use torch.hub instead of local yolov5 repo (avoids ultralytics dependency)
    device = 'cuda' if torch.cuda.is_available() else 'cpu'
    print(f"[INFO] Using device: {device}")
    
    # Load YOLOv5 nano model (fastest, best for Jetson Nano)
    model = torch.hub.load('ultralytics/yolov5', 'yolov5n', pretrained=True)
    model.to(device)
    model.conf = conf_thres
    model.iou = iou_thres
    
    print("[INFO] YOLOv5n model loaded successfully")
except Exception as e:
    print(f"[ERROR] Could not load YOLOv5 model: {e}")
    print("[WARN] Object detection will be DISABLED")
    print("[INFO] To fix: pip3 install torch torchvision")

############################################
# ============ VIDEO CAPTURE ==============
############################################
print(f"[INFO] Attempting to open camera (index: {camera_index})...")
cap = cv2.VideoCapture(camera_index)

if not cap.isOpened():
    print(f"[ERROR] Could not open camera index {camera_index}")
    print("[INFO] Scanning for available cameras...")
    
    available_cameras = []
    for idx in range(5):  # Check first 5 indices
        test_cap = cv2.VideoCapture(idx)
        if test_cap.isOpened():
            available_cameras.append(idx)
            test_cap.release()
    
    if available_cameras:
        print(f"[INFO] Available cameras found at indices: {available_cameras}")
        print("[INFO] On Jetson Nano:")
        print("  - Index 0: Usually CSI camera (built-in)")
        print("  - Index 1+: USB camera")
        print(f"[INFO] Using camera index: {available_cameras[0]} (usually USB)")
        camera_index = available_cameras[0]
        cap = cv2.VideoCapture(camera_index)
    else:
        print("[ERROR] No cameras found! Check:")
        print("  - USB camera is connected")
        print("  - Run: ls /dev/video* to verify")
        exit(1)

print(f"[INFO] Camera opened successfully at index {camera_index}")

# Set camera properties for better performance
cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
cap.set(cv2.CAP_PROP_FPS, 30)

# Initialize motion detection
ret, prev_frame = cap.read()
if not ret:
    print("[ERROR] Failed to read initial frame")
    exit(1)

prev_gray = cv2.cvtColor(prev_frame, cv2.COLOR_BGR2GRAY)
prev_gray = cv2.GaussianBlur(prev_gray, (21, 21), 0)

############################################
# ============ STATE / COOLDOWNS ==========
############################################
last_unknown_time = None
last_motion_time = None
last_object_times = {}
frame_count = 0
fps_start_time = time.time()
fps = 0

############################################
# ============== MAIN LOOP ================
############################################
print("[INFO] Starting security camera system...")
print(f"[INFO] Web stream available at: http://<jetson-ip>:{WEB_PORT}")
print("[INFO] Press Ctrl+C to quit")

# Start Flask web server in background thread
def run_flask():
    try:
        app.run(host='0.0.0.0', port=WEB_PORT, debug=False, use_reloader=False, threaded=True)
    except Exception as e:
        print(f"[ERROR] Flask server error: {e}")

flask_thread = threading.Thread(target=run_flask, daemon=True)
flask_thread.start()
time.sleep(1)  # Wait for Flask to start

print("[INFO] Flask server started on port " + str(WEB_PORT))

try:
    while True:
        ok, frame = cap.read()
        if not ok:
            print("[ERROR] Frame grab failed")
            time.sleep(0.1)
            continue
        
        frame_count += 1
        now = datetime.datetime.now()
        
        # Calculate FPS
        if frame_count % 30 == 0:
            fps = 30 / (time.time() - fps_start_time)
            fps_start_time = time.time()
        
        face_found = False
        object_found = False
        
        # Process only every Nth frame for performance
        if frame_count % PROCESS_EVERY_N_FRAMES == 0:
            
            # ---------- Face Recognition ----------
            try:
                small = cv2.resize(frame, (0, 0), fx=FACE_SCALE_FACTOR, fy=FACE_SCALE_FACTOR)
                rgb_small = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)
                
                face_locs = face_recognition.face_locations(rgb_small, model="hog")  # Use HOG (faster)
                face_encs = face_recognition.face_encodings(rgb_small, face_locs)
                face_found = len(face_locs) > 0
                
                for (top, right, bottom, left), enc in zip(face_locs, face_encs):
                    name = "Unknown"
                    
                    if len(known_face_encodings) > 0:
                        matches = face_recognition.compare_faces(known_face_encodings, enc, tolerance=0.6)
                        dists = face_recognition.face_distance(known_face_encodings, enc)
                        
                        if len(dists) > 0:
                            best_idx = np.argmin(dists)
                            if matches[best_idx]:
                                name = known_face_names[best_idx]
                    
                    # Scale back to original frame
                    scale = int(1 / FACE_SCALE_FACTOR)
                    top, right, bottom, left = [v * scale for v in (top, right, bottom, left)]
                    
                    # Draw rectangle and name
                    color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
                    cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
                    cv2.rectangle(frame, (left, bottom - 35), (right, bottom), color, cv2.FILLED)
                    cv2.putText(frame, name, (left + 6, bottom - 6),
                                cv2.FONT_HERSHEY_DUPLEX, 0.6, (255, 255, 255), 1)
                    
                    # Handle unknown person
                    if name == "Unknown":
                        if (last_unknown_time is None or 
                            (now - last_unknown_time).total_seconds() > cooldown_unknown):
                            
                            crop = frame[max(0, top):bottom, max(0, left):right]
                            fname = os.path.join(unknown_faces_dir, 
                                                f"unknown_{now.strftime('%Y%m%d_%H%M%S')}.jpg")
                            
                            if crop.size > 0:
                                cv2.imwrite(fname, crop)
                                send_telegram_alert("photo", 
                                    ("🚨 Unknown person detected!", fname))
                                print(f"[ALERT] Unknown person detected - saved to {fname}")
                            
                            last_unknown_time = now
                    
                    # Handle known person
                    else:
                        if mark_attendance(name):
                            send_telegram_alert("text", f"✅ {name} - Attendance marked")
            
            except Exception as e:
                print(f"[ERROR] Face recognition error: {e}")
            
            # ---------- YOLOv5 Object Detection ----------
            if model is not None:
                try:
                    results = model(frame)
                    detections = results.pandas().xyxy[0]
                    
                    for _, det in detections.iterrows():
                        label = det['name']
                        conf = det['confidence']
                        x1, y1, x2, y2 = int(det['xmin']), int(det['ymin']), int(det['xmax']), int(det['ymax'])
                        
                        # Draw bounding box
                        color = (255, 165, 0)
                        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                        cv2.putText(frame, f"{label} {conf:.2f}", (x1, y1 - 10),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
                        
                        # Handle non-person objects
                        if label.lower() != "person":
                            object_found = True
                            last_ts = last_object_times.get(label)
                            
                            if (last_ts is None or 
                                (now - last_ts).total_seconds() > cooldown_object):
                                
                                crop = frame[max(0, y1):y2, max(0, x1):x2]
                                out_path = os.path.join("object_alerts",
                                    f"{label}_{now.strftime('%Y%m%d_%H%M%S')}.jpg")
                                
                                if crop.size > 0:
                                    cv2.imwrite(out_path, crop)
                                    
                                    # Determine if animal or object
                                    animals = {"dog", "cat", "cow", "horse", "sheep", "bird", 
                                              "elephant", "bear", "zebra", "giraffe"}
                                    emoji = "🐾" if label.lower() in animals else "📦"
                                    msg = f"{emoji} {label.capitalize()} detected!"
                                    
                                    send_telegram_alert("photo", (msg, out_path))
                                    print(f"[ALERT] {label} detected - saved to {out_path}")
                                
                                last_object_times[label] = now
                
                except Exception as e:
                    print(f"[ERROR] Object detection error: {e}")
            
            # ---------- Motion Detection ----------
            try:
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                gray = cv2.GaussianBlur(gray, (21, 21), 0)
                diff = cv2.absdiff(prev_gray, gray)
                _, thresh = cv2.threshold(diff, 25, 255, cv2.THRESH_BINARY)
                thresh = cv2.dilate(thresh, None, iterations=2)
                contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                
                motion_detected = any(cv2.contourArea(c) > min_motion_area for c in contours)
                
                if motion_detected and not face_found and not object_found:
                    if (last_motion_time is None or 
                        (now - last_motion_time).total_seconds() > cooldown_motion):
                        
                        send_telegram_alert("text", 
                            "⚠️ Motion detected (no person/object identified)")
                        print("[ALERT] Motion detected")
                        last_motion_time = now
                
                prev_gray = gray
            
            except Exception as e:
                print(f"[ERROR] Motion detection error: {e}")
        
        # ---------- Display Frame ----------
        # Add FPS and status overlay
        cv2.putText(frame, f"FPS: {fps:.1f}", (10, 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
        
        status_text = "ACTIVE" if face_found or object_found else "MONITORING"
        status_color = (0, 255, 0) if face_found or object_found else (255, 255, 0)
        cv2.putText(frame, status_text, (10, 60),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, status_color, 2)
        
        # Update frame for web streaming
        with frame_lock:
            current_frame = frame.copy()

except KeyboardInterrupt:
    print("\n[INFO] Interrupted by user")
except Exception as e:
    print(f"[ERROR] Unexpected error: {e}")
    import traceback
    traceback.print_exc()

finally:
    # Cleanup
    print("[INFO] Cleaning up...")
    
    # Stop Telegram worker
    if telegram_enabled:
        telegram_queue.put(None)
        telegram_thread.join(timeout=2)
    
    # Release resources
    cap.release()
    cv2.destroyAllWindows()
    
    print("[INFO] Security camera system stopped")
